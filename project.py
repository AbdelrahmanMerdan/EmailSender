# Python library to read and write in Excel
from openpyxl import Workbook, load_workbook
from email.message import EmailMessage
import win32com.client

# SEND EMAIL
wb = load_workbook('emaillist.xlsx')
ws = wb.active
# number of users range
for row in range(2, 131):
    outlook = win32com.client.Dispatch('outlook.application')

    body = """
    The body of the email goes here.

    """.format(str(ws['B' + str(row)].value), str(ws['A' + str(row)].value))

    for account in outlook.Session.Accounts:
        # Email here
        if account.DisplayName == "email name":

            mail = outlook.CreateItem(0)
            mail.To = str(ws['C' + str(row)].value)

            print(str(ws['C' + str(row)].value))
            mail.Subject = "Email Subject " 
            
            mail.HTMLBody = body
            mail.Importance = 2
            mail.Send()

